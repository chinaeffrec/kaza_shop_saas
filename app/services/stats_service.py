"""Бизнес-логика статистики и экспорта."""
import asyncio
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Optional
from zoneinfo import ZoneInfo

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.schemas.stats import (
    AnalyticsResponse, CohortStats, DashboardResponse,
    ProductStatItem, ProductStatsSummary, ProductStatsResponse,
    RecentOrder, RevenuePoint, StatusStat, StatusStatItem, TopProduct,
)
from app.core.config import get_settings
from app.models.order import ORDER_STATUSES, Order, OrderItem
from app.models.product import Product
from app.models.product_stats import ProductStats
from app.models.user import User


def _fmt_price(price: int | float) -> str:
    return f"{int(price):,} ₽".replace(",", " ")


def _fmt_dt(dt: datetime | None) -> str:
    """Форматирует datetime из UTC в локальную временную зону."""
    if not dt:
        return ""
    tz = ZoneInfo(get_settings().timezone)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(tz).strftime("%d.%m.%Y %H:%M")


async def get_dashboard(
    date_from: Optional[str], date_to: Optional[str],
    session: AsyncSession, shop_id: int = 1,
) -> DashboardResponse:
    # B3: orders и top products запускаются параллельно — они независимы
    orders_q = select(Order).where(Order.shop_id == shop_id).order_by(Order.created_at.desc())
    if date_from:
        orders_q = orders_q.where(Order.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        orders_q = orders_q.where(Order.created_at <= datetime.fromisoformat(date_to + "T23:59:59"))

    # Выручка по топ-товарам за тот же период
    top_revenue_q = (
        select(
            OrderItem.product_id,
            func.sum(OrderItem.quantity * OrderItem.price).label("revenue"),
        )
        .join(Order, OrderItem.order_id == Order.id)
        .where(Order.shop_id == shop_id, Order.status.not_in(["cancelled", "returned"]))
        .group_by(OrderItem.product_id)
        .order_by(func.sum(OrderItem.quantity * OrderItem.price).desc())
        .limit(10)
    )
    if date_from:
        top_revenue_q = top_revenue_q.where(Order.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        top_revenue_q = top_revenue_q.where(
            Order.created_at <= datetime.fromisoformat(date_to + "T23:59:59"))

    top_q = (
        select(ProductStats, Product)
        .outerjoin(Product, ProductStats.product_id == Product.id)
        .where(Product.shop_id == shop_id)
        .order_by(ProductStats.ordered.desc()).limit(10)
    )

    orders_result, top_result, top_rev_result = await asyncio.gather(
        session.execute(orders_q),
        session.execute(top_q),
        session.execute(top_revenue_q),
    )
    all_orders = orders_result.scalars().all()
    top_rows = top_result.all()
    top_revenue_map: dict[int, int] = {
        r.product_id: int(r.revenue or 0) for r in top_rev_result.all()
    }

    REVENUE_STATUSES = {"paid", "confirmed", "assembled", "shipped", "delivered"}
    billable = [o for o in all_orders if o.status in REVENUE_STATUSES]
    total_revenue = sum(o.total for o in billable)
    avg_val = int(total_revenue / len(billable)) if billable else 0

    by_status: dict[str, int] = {}
    for o in all_orders:
        by_status[o.status] = by_status.get(o.status, 0) + 1

    # user_map зависит от all_orders — запускаем после
    user_ids = {o.user_id for o in all_orders}
    user_map: dict = {}
    if user_ids:
        users = (await session.execute(
            select(User).where(User.telegram_id.in_(user_ids), User.shop_id == shop_id)
        )).scalars().all()
        user_map = {u.telegram_id: u for u in users}

    recent = []
    for o in all_orders[:10]:
        u = user_map.get(o.user_id)
        name = f"ID:{o.user_id}"
        if u:
            name = f"{u.first_name or ''} {u.last_name or ''}".strip() or u.username or name
        recent.append(RecentOrder(
            id=o.id, user_id=o.user_id, user_name=name,
            status=o.status, status_label=ORDER_STATUSES.get(o.status, o.status),
            total=o.total, created_at=o.created_at.isoformat(),
        ))

    top = [TopProduct(
        product_id=ps.product_id,
        name=p.name if p else f"ID:{ps.product_id}",
        ordered=ps.ordered or 0,
        revenue=top_revenue_map.get(ps.product_id, 0),
    ) for ps, p in top_rows]

    status_order = list(ORDER_STATUSES.keys())
    sorted_statuses = [s for s in status_order if s in by_status] + \
                      [s for s in by_status if s not in status_order]

    total_orders_count = len(all_orders)
    conversion_rate = round(len(billable) / total_orders_count * 100, 1) if total_orders_count else 0.0

    return DashboardResponse(
        total_revenue=total_revenue,
        total_orders=total_orders_count,
        billable_orders=len(billable),
        average_order_value=avg_val,
        conversion_rate=conversion_rate,
        by_status={s: StatusStat(count=by_status[s], label=ORDER_STATUSES.get(s, s))
                   for s in sorted_statuses},
        orders_by_status=[
            StatusStatItem(status=s, count=by_status[s], label=ORDER_STATUSES.get(s, s))
            for s in sorted_statuses
        ],
        recent_orders=recent, top_products=top,
    )


async def export_orders_xlsx(
    date_from: Optional[str], date_to: Optional[str],
    status: Optional[str], session: AsyncSession, shop_id: int = 1,
) -> BytesIO:
    q = select(Order).where(Order.shop_id == shop_id).order_by(Order.created_at.desc())
    if date_from:
        q = q.where(Order.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.where(Order.created_at <= datetime.fromisoformat(date_to + "T23:59:59"))
    if status:
        q = q.where(Order.status == status)

    orders = (await session.execute(q)).scalars().all()
    order_ids = [o.id for o in orders]

    user_ids = {o.user_id for o in orders}
    user_map: dict = {}
    if user_ids:
        users = (await session.execute(
            select(User).where(User.telegram_id.in_(user_ids), User.shop_id == shop_id)
        )).scalars().all()
        user_map = {u.telegram_id: u for u in users}

    # Состав заказов — одним запросом
    items_map: dict[int, list] = {}
    if order_ids:
        for item in (await session.execute(
            select(OrderItem).where(OrderItem.order_id.in_(order_ids))
        )).scalars().all():
            items_map.setdefault(item.order_id, []).append(item)

    # B4: Excel-генерация (синхронная, CPU-bound) выносится из event loop
    def _build_xlsx() -> BytesIO:
        wb = Workbook()
        hfill = PatternFill("solid", fgColor="4472C4")
        hfont = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _hdr(ws, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = hfill; cell.font = hfont; cell.border = border

        # ── Лист 1: Заказы ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Заказы"
        _hdr(ws1, ["ID", "Покупатель", "Контакт", "Телефон",
                   "Сумма", "Статус", "Адрес", "Комментарий", "Заметки", "Дата"])

        for ri, order in enumerate(orders, 2):
            u = user_map.get(order.user_id)
            uname = f"ID:{order.user_id}"
            ucontact = uphone = ""
            if u:
                uname = (f"{u.first_name or ''} {u.last_name or ''}".strip()
                         or u.username or uname)
                ucontact = f"@{u.username}" if u.username else ""
                uphone = getattr(u, "phone", "") or ""
            row_data = [
                order.id, uname, ucontact, uphone, order.total,
                ORDER_STATUSES.get(order.status, order.status),
                getattr(order, "delivery_address", "") or "",
                order.comment or "",
                getattr(order, "staff_notes", "") or "",
                _fmt_dt(order.created_at),
            ]
            for col, val in enumerate(row_data, 1):
                ws1.cell(row=ri, column=col, value=val).border = border

        for i, w in enumerate([8, 28, 18, 16, 12, 16, 30, 30, 30, 18], 1):
            ws1.column_dimensions[ws1.cell(1, i).column_letter].width = w

        # ── Лист 2: Состав заказов ──────────────────────────────────────
        ws2 = wb.create_sheet("Состав заказов")
        _hdr(ws2, ["ID заказа", "Товар", "Вариант", "SKU", "Цена", "Кол-во", "Сумма"])

        ri2 = 2
        for order in orders:
            for item in items_map.get(order.id, []):
                for col, val in enumerate([
                    order.id, item.name,
                    getattr(item, "variant_name", "") or "",
                    getattr(item, "sku", "") or "",
                    item.price, item.quantity, item.price * item.quantity,
                ], 1):
                    ws2.cell(row=ri2, column=col, value=val).border = border
                ri2 += 1

        for i, w in enumerate([12, 40, 20, 16, 12, 10, 12], 1):
            ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    return await asyncio.to_thread(_build_xlsx)


def _build_period_filter(q, date_from: Optional[str], date_to: Optional[str]):
    if date_from:
        q = q.where(Order.created_at >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.where(Order.created_at <= datetime.fromisoformat(date_to + "T23:59:59"))
    return q


async def get_product_stats(
    date_from: Optional[str], date_to: Optional[str],
    session: AsyncSession, shop_id: int = 1,
) -> ProductStatsResponse:
    ps_res = await session.execute(
        select(ProductStats, Product)
        .outerjoin(Product, ProductStats.product_id == Product.id)
        .where(Product.shop_id == shop_id)
        .order_by(ProductStats.ordered.desc())
    )
    rows = ps_res.all()

    oi_q = (
        select(
            OrderItem.product_id,
            func.sum(OrderItem.quantity).label("qty"),
            func.sum(OrderItem.quantity * OrderItem.price).label("revenue"),
        )
        .join(Order, OrderItem.order_id == Order.id)
        .where(Order.shop_id == shop_id)
        .where(Order.status.not_in(["cancelled", "returned"]))
        .group_by(OrderItem.product_id)
    )
    oi_q = _build_period_filter(oi_q, date_from, date_to)
    period_map: dict[int, tuple[int, int]] = {
        row.product_id: (int(row.qty), int(row.revenue))
        for row in (await session.execute(oi_q)).all()
    }

    order_count_q = (
        select(func.count(func.distinct(OrderItem.order_id)))
        .join(Order, OrderItem.order_id == Order.id)
        .where(Order.shop_id == shop_id)
        .where(Order.status.not_in(["cancelled", "returned"]))
    )
    order_count_q = _build_period_filter(order_count_q, date_from, date_to)
    order_count = (await session.execute(order_count_q)).scalar() or 0

    items: list[ProductStatItem] = []
    total_sold_qty = total_sold_sum = total_returned = 0

    for ps, product in rows:
        qty, revenue = period_map.get(ps.product_id, (0, 0))
        total_sold_qty += qty
        total_sold_sum += revenue
        total_returned += ps.returned or 0
        items.append(ProductStatItem(
            product_id=ps.product_id,
            name=product.name if product else f"ID:{ps.product_id}",
            added_to_cart=ps.added_to_cart or 0,
            ordered=ps.ordered or 0,
            returned=ps.returned or 0,
            period_sold_qty=qty,
            period_sold_sum=revenue,
        ))

    avg_price = int(total_sold_sum / total_sold_qty) if total_sold_qty else 0

    summary = ProductStatsSummary(
        total_sold_qty=total_sold_qty,
        total_sold_sum=total_sold_sum,
        avg_items_per_order=round(total_sold_qty / order_count, 1) if order_count else 0,
        avg_price=avg_price,
        total_returned=total_returned,
    )
    return ProductStatsResponse(items=items, summary=summary)


async def export_product_stats_xlsx(
    date_from: Optional[str], date_to: Optional[str],
    session: AsyncSession, shop_id: int = 1,
) -> BytesIO:
    result = await get_product_stats(date_from, date_to, session, shop_id)

    # B4: Excel-генерация выносится из event loop
    def _build_xlsx() -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика товаров"
        hfill = PatternFill("solid", fgColor="4472C4")
        hfont = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        hdrs = ["Товар", "В корзину", "Заказано (всего)", "Возвраты", "Продано (период)", "Выручка (период)"]
        for col, h in enumerate(hdrs, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill; cell.font = hfont; cell.border = border

        for ri, item in enumerate(result.items, 2):
            for col, val in enumerate([
                item.name, item.added_to_cart, item.ordered,
                item.returned, item.period_sold_qty, item.period_sold_sum,
            ], 1):
                ws.cell(row=ri, column=col, value=val).border = border

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    return await asyncio.to_thread(_build_xlsx)


async def track_return(product_id: int, session: AsyncSession) -> dict:
    res = await session.execute(
        select(ProductStats).where(ProductStats.product_id == product_id)
    )
    ps = res.scalar_one_or_none()
    if not ps:
        raise HTTPException(404, "Product stats not found")
    ps.returned = (ps.returned or 0) + 1
    await session.commit()
    return {"ok": True, "product_id": product_id, "returned": ps.returned}


# ── Analytics: revenue chart + cohorts ───────────────────────────────────────

_BILLABLE_STATUSES = {"paid", "confirmed", "assembled", "shipped", "delivered"}


def _default_date_range() -> tuple[datetime, datetime]:
    """Последние 30 дней если период не задан."""
    now = datetime.now(timezone.utc)
    return now - timedelta(days=29), now


async def get_analytics(
    date_from: Optional[str],
    date_to: Optional[str],
    period: str,          # "day" | "week"
    session: AsyncSession,
    shop_id: int,
) -> AnalyticsResponse:
    """
    Возвращает:
    - revenue_chart: динамика выручки по дням или неделям
    - cohorts: новые vs. повторные покупатели
    """
    if period not in ("day", "week"):
        period = "day"

    dt_from, dt_to = _default_date_range()
    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from).replace(
                hour=0, minute=0, second=0, tzinfo=timezone.utc
            )
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to).replace(
                hour=23, minute=59, second=59, tzinfo=timezone.utc
            )
        except ValueError:
            pass

    # ── 1. Revenue chart ──────────────────────────────────────────────────────
    trunc_expr = func.date_trunc(period, Order.created_at)
    chart_q = (
        select(
            trunc_expr.label("bucket"),
            func.sum(Order.total).label("revenue"),
            func.count(Order.id).label("orders_count"),
        )
        .where(
            Order.shop_id == shop_id,
            Order.status.not_in(["cancelled", "returned"]),
            Order.created_at >= dt_from,
            Order.created_at <= dt_to,
        )
        .group_by(trunc_expr)
        .order_by(trunc_expr)
    )

    # Новые покупатели по бакетам (первый заказ именно в этом бакете)
    first_order_sub = (
        select(
            Order.user_id,
            func.min(Order.created_at).label("first_at"),
        )
        .where(Order.shop_id == shop_id)
        .group_by(Order.user_id)
        .subquery()
    )
    new_cust_q = (
        select(
            func.date_trunc(period, first_order_sub.c.first_at).label("bucket"),
            func.count(first_order_sub.c.user_id).label("new_count"),
        )
        .where(
            first_order_sub.c.first_at >= dt_from,
            first_order_sub.c.first_at <= dt_to,
        )
        .group_by(func.date_trunc(period, first_order_sub.c.first_at))
    )

    chart_rows, new_cust_rows = await asyncio.gather(
        session.execute(chart_q),
        session.execute(new_cust_q),
    )

    new_cust_map: dict[str, int] = {}
    for r in new_cust_rows.all():
        if r.bucket:
            key = r.bucket.strftime("%Y-W%W" if period == "week" else "%Y-%m-%d")
            new_cust_map[key] = int(r.new_count or 0)

    revenue_chart: list[RevenuePoint] = []
    for r in chart_rows.all():
        if not r.bucket:
            continue
        label = r.bucket.strftime("%Y-W%W" if period == "week" else "%Y-%m-%d")
        revenue_chart.append(RevenuePoint(
            date=label,
            revenue=int(r.revenue or 0),
            orders_count=int(r.orders_count or 0),
            new_customers=new_cust_map.get(label, 0),
        ))

    # ── 2. Cohort analysis ────────────────────────────────────────────────────
    # Покупатели, у которых есть хотя бы один заказ в периоде
    period_buyers_q = (
        select(func.distinct(Order.user_id))
        .where(Order.shop_id == shop_id, Order.created_at >= dt_from, Order.created_at <= dt_to)
    )
    period_buyer_ids = [r[0] for r in (await session.execute(period_buyers_q)).all()]

    new_buyers = returning_buyers = 0
    avg_orders_new = avg_orders_returning = 0.0

    if period_buyer_ids:
        # Покупатели с заказами ДО периода → повторные
        before_q = (
            select(func.distinct(Order.user_id))
            .where(
                Order.shop_id == shop_id,
                Order.user_id.in_(period_buyer_ids),
                Order.created_at < dt_from,
            )
        )
        returning_ids = {r[0] for r in (await session.execute(before_q)).all()}
        new_ids = set(period_buyer_ids) - returning_ids

        new_buyers = len(new_ids)
        returning_buyers = len(returning_ids)

        # Среднее количество заказов в периоде у каждой когорты
        if new_ids or returning_ids:
            orders_per_user_q = (
                select(
                    Order.user_id,
                    func.count(Order.id).label("cnt"),
                )
                .where(
                    Order.shop_id == shop_id,
                    Order.created_at >= dt_from, Order.created_at <= dt_to,
                    Order.user_id.in_(period_buyer_ids),
                )
                .group_by(Order.user_id)
            )
            opu_rows = (await session.execute(orders_per_user_q)).all()
            opu_map = {r.user_id: int(r.cnt) for r in opu_rows}

            if new_ids:
                avg_orders_new = round(
                    sum(opu_map.get(uid, 0) for uid in new_ids) / len(new_ids), 2
                )
            if returning_ids:
                avg_orders_returning = round(
                    sum(opu_map.get(uid, 0) for uid in returning_ids) / len(returning_ids), 2
                )

    period_count = len(period_buyer_ids)
    repeat_rate = round(returning_buyers / period_count * 100, 1) if period_count else 0.0

    cohorts = CohortStats(
        period_buyers=period_count,
        new_buyers=new_buyers,
        returning_buyers=returning_buyers,
        repeat_rate=repeat_rate,
        avg_orders_new=avg_orders_new,
        avg_orders_returning=avg_orders_returning,
    )

    return AnalyticsResponse(
        revenue_chart=revenue_chart,
        cohorts=cohorts,
        period=period,
    )


# ── Customer export ────────────────────────────────────────────────────────────

async def export_customers_xlsx(session: AsyncSession, shop_id: int) -> BytesIO:
    """
    Экспортирует всех покупателей с LTV-агрегатами в Excel.
    Данные: имя, username, телефон, заказов, выручка, первый/последний заказ, заметки.
    """
    from app.services.customer_service import list_customers as _list_customers

    # Грузим всех без пагинации (limit = 10 000 — защита от OOM)
    items, _ = await _list_customers(page=1, per_page=10_000, session=session, shop_id=shop_id)

    def _build() -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Покупатели"

        hfill = PatternFill("solid", fgColor="4472C4")
        hfont = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "Telegram ID", "Имя", "Фамилия", "@username", "Телефон",
            "Заказов", "Потрачено (₽)", "Первый заказ", "Последний заказ",
            "Заметки",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill; cell.font = hfont; cell.border = border
            cell.alignment = Alignment(horizontal="center")

        def _d(dt) -> str:
            return dt.strftime("%d.%m.%Y") if dt else ""

        for ri, c in enumerate(items, 2):
            row = [
                c.telegram_id,
                c.first_name or "",
                c.last_name or "",
                f"@{c.username}" if c.username else "",
                c.phone or "",
                c.total_orders,
                c.total_spent,
                _d(c.first_order_at),
                _d(c.last_order_at),
                c.notes or "",
            ]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=col, value=val)
                cell.border = border

        col_widths = [14, 18, 18, 18, 16, 10, 16, 14, 14, 40]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    return await asyncio.to_thread(_build)
